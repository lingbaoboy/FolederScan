import tkinter as tk
from tkinter import filedialog, messagebox
import threading
from pathlib import Path
from datetime import datetime
import subprocess
import sys
import os
import openpyxl
from openpyxl.utils import get_column_letter


class DirectoryScannerApp:
    def __init__(self, root):
        """初始化应用程序UI和变量"""
        self.root = root
        self.root.title("目录扫描工具(XLSX输出版)")
        self.root.geometry("600x520")
        self.root.minsize(550, 520)

        # UI 变量和布局与 v6.0 完全相同
        self.start_dir = tk.StringVar()
        self.depth_level = tk.StringVar(value="3")
        self.scan_all_subfolders = tk.BooleanVar(value=False)
        self.stop_keyword = tk.StringVar(value=".git node_modules")
        self.stop_keyword_case_sensitive = tk.BooleanVar(value=False)
        self.exclude_types = tk.StringVar(value=".log .tmp .bak")
        self.exclude_types_case_sensitive = tk.BooleanVar(value=False)
        self.filename_keyword = tk.StringVar()
        self.filename_keyword_case_sensitive = tk.BooleanVar(value=False)
        self.filename_keyword_mode = tk.StringVar(value="blacklist")
        main_frame = tk.Frame(root, padx=15, pady=15)
        main_frame.pack(fill=tk.BOTH, expand=True)
        main_frame.columnconfigure(0, weight=1)
        dir_frame = tk.Frame(main_frame)
        dir_frame.grid(row=0, column=0, sticky="ew", pady=5)
        dir_frame.columnconfigure(1, weight=1)
        dir_label = tk.Label(dir_frame, text="起始目录:")
        dir_label.grid(row=0, column=0, padx=(0, 10))
        self.dir_entry = tk.Entry(dir_frame, textvariable=self.start_dir, state='readonly')
        self.dir_entry.grid(row=0, column=1, sticky="ew")
        dir_button = tk.Button(dir_frame, text="选择...", command=self.select_directory)
        dir_button.grid(row=0, column=2, padx=(10, 0))
        options_frame = tk.LabelFrame(main_frame, text="扫描参数", padx=10, pady=10)
        options_frame.grid(row=1, column=0, sticky="ew", pady=(10, 0))
        options_frame.columnconfigure(1, weight=1)
        depth_label = tk.Label(options_frame, text="遍历层级:")
        depth_label.grid(row=0, column=0, sticky="w", padx=(0, 10), pady=2)
        self.depth_entry = tk.Entry(options_frame, textvariable=self.depth_level, width=10)
        self.depth_entry.grid(row=0, column=1, sticky="w", pady=2)
        self.scan_all_check = tk.Checkbutton(options_frame, text="遍历所有子文件夹", variable=self.scan_all_subfolders,
                                             command=self._toggle_depth_entry_state)
        self.scan_all_check.grid(row=0, column=2, sticky="w", padx=(10, 0))
        fname_label = tk.Label(options_frame, text="文件名关键词:")
        fname_label.grid(row=1, column=0, sticky="w", padx=(0, 10), pady=(8, 2))
        fname_entry = tk.Entry(options_frame, textvariable=self.filename_keyword)
        fname_entry.grid(row=1, column=1, sticky="we", pady=(8, 2))
        fname_case_check = tk.Checkbutton(options_frame, text="大小写敏感",
                                          variable=self.filename_keyword_case_sensitive)
        fname_case_check.grid(row=1, column=2, sticky="w", padx=(10, 0))
        fname_mode_frame = tk.Frame(options_frame)
        fname_mode_frame.grid(row=2, column=1, columnspan=2, sticky="w")
        tk.Radiobutton(fname_mode_frame, text="忽略含关键词项 (黑名单)", variable=self.filename_keyword_mode,
                       value="blacklist").pack(side=tk.LEFT)
        tk.Radiobutton(fname_mode_frame, text="仅保留含关键词项 (白名单)", variable=self.filename_keyword_mode,
                       value="whitelist").pack(side=tk.LEFT, padx=10)
        stop_label = tk.Label(options_frame, text="目录停止关键字:")
        stop_label.grid(row=3, column=0, sticky="w", padx=(0, 10), pady=(8, 2))
        stop_entry = tk.Entry(options_frame, textvariable=self.stop_keyword)
        stop_entry.grid(row=3, column=1, sticky="we", pady=(8, 2))
        stop_case_check = tk.Checkbutton(options_frame, text="大小写敏感", variable=self.stop_keyword_case_sensitive)
        stop_case_check.grid(row=3, column=2, sticky="w", padx=(10, 0))
        stop_note = tk.Label(options_frame, text="多个关键字用空格分隔", fg="grey")
        stop_note.grid(row=4, column=1, sticky="w")
        exclude_label = tk.Label(options_frame, text="排除文件类型:")
        exclude_label.grid(row=5, column=0, sticky="w", padx=(0, 10), pady=(8, 2))
        exclude_entry = tk.Entry(options_frame, textvariable=self.exclude_types)
        exclude_entry.grid(row=5, column=1, sticky="we", pady=(8, 2))
        exclude_case_check = tk.Checkbutton(options_frame, text="大小写敏感",
                                            variable=self.exclude_types_case_sensitive)
        exclude_case_check.grid(row=5, column=2, sticky="w", padx=(10, 0))
        exclude_note = tk.Label(options_frame, text="多个类型用空格分隔, 如: .pdf .docx", fg="grey")
        exclude_note.grid(row=6, column=1, sticky="w")
        self.start_button = tk.Button(main_frame, text="开始扫描", command=self.start_scan_thread,
                                      font=("", 12, "bold"))
        self.start_button.grid(row=2, column=0, pady=(20, 10))
        self.status_frame = tk.Frame(main_frame, bd=1, relief=tk.SUNKEN)
        self.status_frame.grid(row=3, column=0, sticky="ew", ipady=5)
        self._update_status_display("请先选择起始目录和设置参数")

    def run_scan(self, start_path_str, max_depth, stop_keyword_set, stop_case_sensitive, exclude_set,
                 exclude_case_sensitive, fname_keyword_set, fname_case_sensitive, fname_mode):
        """修改：执行目录遍历和 XLSX 文件写入的核心逻辑"""
        try:
            start_path = Path(start_path_str)
            output_file_path = start_path / "scan_results.xlsx"  # 输出文件名更改为 .xlsx

            # 使用 openpyxl 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "扫描结果"
            ws.append(["文件名", "相对路径", "修改时间", "创建时间"])

            # 递归扫描，传入 worksheet 对象
            self._recursive_scan(start_path, start_path, 0, max_depth,
                                 stop_keyword_set, stop_case_sensitive,
                                 exclude_set, exclude_case_sensitive,
                                 fname_keyword_set, fname_case_sensitive, fname_mode,
                                 ws)  # 传入 worksheet

            # 调整列宽并保存
            self.adjust_column_width(ws)
            wb.save(output_file_path)

            self.root.after(0, self.on_scan_complete, output_file_path)
        except Exception as e:
            self.root.after(0, self.on_scan_error, e)
        finally:
            self.root.after(0, lambda: self.start_button.config(state=tk.NORMAL, text="开始扫描"))

    def _recursive_scan(self, current_path, base_path, current_depth, max_depth,
                        stop_keyword_set, stop_case_sensitive,
                        exclude_set, exclude_case_sensitive,
                        fname_keyword_set, fname_case_sensitive, fname_mode,
                        worksheet):  # 修改：接收 worksheet
        """核心修正：将“写入日志”和“深入遍历”的逻辑彻底分离"""
        if current_depth >= max_depth: return
        try:
            for item in current_path.iterdir():
                try:
                    # 决策1: 判断是否应将此项目写入日志
                    if self._should_log_item(item, exclude_set, exclude_case_sensitive, fname_keyword_set,
                                             fname_case_sensitive, fname_mode):
                        mod_time = datetime.fromtimestamp(item.stat().st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                        create_time = datetime.fromtimestamp(item.stat().st_ctime).strftime('%Y-%m-%d %H:%M:%S')
                        relative_path = item.relative_to(base_path).as_posix()

                        # 修改：将数据行追加到 worksheet
                        data_row = [item.name, relative_path, mod_time, create_time]
                        worksheet.append(data_row)

                    # 决策 2: 判断是否应深入遍历此目录
                    if item.is_dir():
                        name_to_check = item.name if stop_case_sensitive else item.name.lower()
                        if stop_keyword_set and any(keyword in name_to_check for keyword in stop_keyword_set):
                            continue

                        # 修正：在递归调用中传递 worksheet
                        self._recursive_scan(item, base_path, current_depth + 1, max_depth,
                                             stop_keyword_set, stop_case_sensitive,
                                             exclude_set, exclude_case_sensitive,
                                             fname_keyword_set, fname_case_sensitive, fname_mode,
                                             worksheet)  # 传递 worksheet
                except (OSError, PermissionError) as e:
                    print(f"无法访问 {item}: {e}")
                except Exception as e:
                    print(f"处理 {item} 时发生未知错误: {e}")
        except (OSError, PermissionError) as e:
            print(f"无法访问目录 {current_path}: {e}")

    def adjust_column_width(self, worksheet):
        """新增：自动调整Excel列宽"""
        for col in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            worksheet.column_dimensions[column_letter].width = max_length + 2

    # --- 其他所有方法与 v6.0 版本完全相同 ---
    def _should_log_item(self, item, exclude_set, exclude_case_sensitive, fname_keyword_set, fname_case_sensitive,
                         fname_mode):
        if item.is_file():
            suffix_to_check = item.suffix if exclude_case_sensitive else item.suffix.lower()
            if suffix_to_check in exclude_set: return False
        if fname_keyword_set:
            name_to_check = item.name if fname_case_sensitive else item.name.lower()
            is_matched = any(keyword in name_to_check for keyword in fname_keyword_set)
            if fname_mode == 'blacklist' and is_matched: return False
            if fname_mode == 'whitelist' and not is_matched: return False
        return True

    def start_scan_thread(self):
        start_path_str = self.start_dir.get()
        if not start_path_str: messagebox.showerror("错误", "请选择一个起始目录！"); return
        if self.scan_all_subfolders.get():
            max_depth = float('inf')
        else:
            try:
                max_depth = int(self.depth_level.get())
                if max_depth <= 0: raise ValueError
            except ValueError:
                messagebox.showerror("错误", "遍历层级必须是一个正整数！"); return
        stop_case_sensitive, exclude_case_sensitive, fname_case_sensitive = self.stop_keyword_case_sensitive.get(), self.exclude_types_case_sensitive.get(), self.filename_keyword_case_sensitive.get()
        fname_mode = self.filename_keyword_mode.get()
        stop_keyword_str = self.stop_keyword.get() if stop_case_sensitive else self.stop_keyword.get().lower()
        exclude_types_str = self.exclude_types.get() if exclude_case_sensitive else self.exclude_types.get().lower()
        fname_keyword_str = self.filename_keyword.get() if fname_case_sensitive else self.filename_keyword.get().lower()
        stop_keyword_set = {kw.strip() for kw in stop_keyword_str.split() if kw.strip()}
        exclude_set = {f".{ext.strip().lstrip('.')}" for ext in exclude_types_str.split() if ext.strip()}
        fname_keyword_set = {kw.strip() for kw in fname_keyword_str.split() if kw.strip()}
        self.start_button.config(state=tk.DISABLED, text="正在扫描...")
        self._update_status_display("扫描已开始...")
        threading.Thread(target=self.run_scan,
                         args=(start_path_str, max_depth, stop_keyword_set, stop_case_sensitive, exclude_set,
                               exclude_case_sensitive, fname_keyword_set, fname_case_sensitive, fname_mode),
                         daemon=True).start()

    def _toggle_depth_entry_state(self):
        if self.scan_all_subfolders.get():
            self.depth_entry.config(state='disabled')
        else:
            self.depth_entry.config(state='normal')

    def _update_status_display(self, message, file_path=None):
        for widget in self.status_frame.winfo_children(): widget.destroy()
        self.status_frame.columnconfigure(0, weight=1)
        if file_path:
            status_label = tk.Label(self.status_frame, text=message, anchor='w')
            status_label.grid(row=0, column=0, sticky='w', padx=(5, 5))
            path_entry = tk.Entry(self.status_frame)
            path_entry.insert(0, str(file_path))
            path_entry.config(state='readonly')
            path_entry.grid(row=1, column=0, sticky='ew', padx=(5, 5))
            open_btn = tk.Button(self.status_frame, text="打开结果", command=lambda: self._open_path(file_path))
            open_btn.grid(row=0, rowspan=2, column=1, sticky='e', padx=(5, 5))
        else:
            status_label = tk.Label(self.status_frame, text=message, anchor='w')
            status_label.grid(row=0, column=0, sticky='ew', padx=(5, 0))

    def select_directory(self):
        path = filedialog.askdirectory()
        if path: self.start_dir.set(path); self._update_status_display(f"已选择目录: {path}")

    def _open_path(self, path):
        try:
            if sys.platform == "win32":
                os.startfile(path)
            elif sys.platform == "darwin":
                subprocess.run(['open', path], check=True)
            else:
                subprocess.run(['xdg-open', path], check=True)
        except (FileNotFoundError, AttributeError, subprocess.CalledProcessError) as e:
            messagebox.showwarning("无法打开文件", f"无法自动打开结果文件。\n请手动导航至:\n{path}\n\n错误: {e}")

    def on_scan_complete(self, output_file_path):
        self._update_status_display("扫描完成！结果已保存至:", output_file_path)

    def on_scan_error(self, error):
        self._update_status_display(f"扫描出错: {error}")


if __name__ == "__main__":
    root = tk.Tk()
    app = DirectoryScannerApp(root)
    root.mainloop()